import os
import logging
import tempfile
import uuid
from typing import Optional

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import formulas
except ImportError:
    formulas = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from pdf2image import convert_from_path
    import pytesseract
    from PIL import Image
except ImportError:
    convert_from_path = None
    pytesseract = None
    Image = None

logger = logging.getLogger(__name__)

def extract_text_from_file(file_path: str, mime_type: Optional[str] = None) -> str:
    """
    Extract text content from a file (PDF, Excel, Image).
    """
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return ""

    ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if ext == '.pdf':
            return _extract_from_pdf(file_path)
        elif ext in ['.xlsx', '.xls', '.csv']:
            return _extract_from_excel(file_path)
        elif ext in ['.txt', '.md']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        else:
            # Note: Images are now handled by Vision API in the main file
            logger.warning(f"File type {ext} routed to text parser instead of Vision.")
            return ""
            
    except Exception as e:
        logger.error(f"Error extracting text from {file_path}: {e}")
        return ""

def _extract_from_pdf(file_path: str) -> str:
    text = ""
    try:
        if PyPDF2:
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
    except Exception as e:
        logger.warning(f"PyPDF2 extraction failed: {e}")

    # Fallback for Scanned PDFs
    if len(text.strip()) < 50:
        logger.info("PDF text empty or too short. Attempting OCR...")
        try:
            if convert_from_path and pytesseract:
                images = convert_from_path(file_path, first_page=1, last_page=3) 
                ocr_text = ""
                for img in images:
                    ocr_text += pytesseract.image_to_string(img) + "\n"
                
                if len(ocr_text.strip()) > len(text.strip()):
                    text = ocr_text
        except Exception as e:
             logger.error(f"OCR failed: {e}")
             
    return text
def _extract_from_excel(file_path: str) -> str:
    text = ""
    ext = os.path.splitext(file_path)[1].lower()

    # 1. Force Formula Calculation (The Headless Excel Fix)
    if ext == '.xlsx' and formulas and openpyxl:
        try:
            logger.info("Attempting to calculate Excel formulas programmatically...")
            xl_model = formulas.ExcelModel().loads(file_path).finish()
            xl_model.calculate()
            
            # Create an actual temporary DIRECTORY instead of a file
            with tempfile.TemporaryDirectory() as temp_dir:
                # Tell formulas to save the output file inside this new folder
                xl_model.write(dirpath=temp_dir)
                
                # Find the newly generated .xlsx file inside the temp folder
                calculated_file = None
                for root, _, files in os.walk(temp_dir):
                    for file in files:
                        if file.lower().endswith('.xlsx'):
                            calculated_file = os.path.join(root, file)
                            break
                    if calculated_file:
                        break
                        
                if not calculated_file:
                    raise FileNotFoundError("Calculated Excel file not found in temp directory.")
                
                # Read the mathematically calculated file
                wb = openpyxl.load_workbook(calculated_file, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text += f"\n--- Sheet: {sheet_name} ---\n"
                    for row in sheet.iter_rows(values_only=True):
                        row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                        if row_values:
                            text += " | ".join(row_values) + "\n"
                            
            # The 'with TemporaryDirectory()' block automatically deletes the folder and files for us!
            return text
        except Exception as e:
            logger.warning(f"Formula calculation failed, falling back to basic extraction: {e}")

    # 2. Fallback to basic openpyxl
    if ext == '.xlsx' and openpyxl:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if row_values:
                        text += " | ".join(row_values) + "\n"
            return text
        except Exception as e:
            logger.warning(f"Fallback openpyxl extraction failed: {e}")

    # 3. Final Fallback to pandas
    if pd:
        try:
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                text += f"--- Sheet: {sheet_name} ---\n"
                text += df.to_string(index=False, na_rep="") + "\n"
        except Exception as e:
            logger.error(f"Pandas extraction failed: {e}")
            
    return text