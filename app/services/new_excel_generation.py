import os
import logging
import shutil
from datetime import datetime
from typing import Dict, Optional

# --- Import OpenPyXL (Works on Linux/Render) ---
try:
    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    raise ImportError("CRITICAL: 'openpyxl' is missing. Add 'openpyxl>=3.1.0' to requirements.txt")

logger = logging.getLogger(__name__)

class ExcelGenerationService:
    """
    Service class for generating quotation Excel files.
    Optimized for Linux/Render (Docker) environments.
    """

    def __init__(
        self,
        template_path: str = "sample/QuotationFormat.xlsx",
        output_dir: str = "generated",
    ):
        self.template_path = template_path
        self.output_dir = output_dir

        # Ensure output directory exists - safer cleanup for Docker to avoid "Device Busy"
        if os.path.exists(output_dir):
            for filename in os.listdir(output_dir):
                file_path = os.path.join(output_dir, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    logger.error(f"Failed to delete {file_path}. Reason: {e}")
        
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Excel Service initialized. Template: {template_path}")

    def generate_quotation_excel(
        self, gmail_id: str, extraction_data: Dict, copy_only: bool = False
    ) -> Optional[str]:
        """
        Generate a quotation Excel file using OpenPyXL (Render Compatible).
        """
        try:
            # 1. Validation
            if not os.path.exists(self.template_path):
                logger.error(f"❌ Template not found at: {self.template_path}")
                return None

            # 2. Setup Filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"quotation_{gmail_id}_{timestamp}.xlsx"
            output_path = os.path.join(self.output_dir, filename)

            # 3. Copy Template
            shutil.copy2(self.template_path, output_path)

            if copy_only:
                return output_path

            # 4. Fill Data
            extraction_result = extraction_data.get("extraction_result", {})
            
            # Using OpenPyXL to edit the file
            wb = load_workbook(output_path)
            ws = wb.active
            
            self._fill_data(ws, extraction_result)
            
            wb.save(output_path)
            wb.close()
            
            logger.info(f"✅ Excel generated successfully: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def _fill_data(self, ws, extraction_result):
        """
        Fills the Excel data with EXACT formatting, Borders, and Colors.
        """
        requirements = extraction_result.get("Requirements", [])
        
        # --- STYLING CONSTANTS ---
        # 1. Borders
        thin_side = Side(border_style="thin", color="000000")
        full_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
        
        thick_side = Side(border_style="medium", color="000000")
        thick_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
        
        # 2. Alignments
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_top_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")
        left_center_align = Alignment(horizontal="left", vertical="center")
        
        # 3. Fonts
        red_bold_font = Font(name='Calibri', size=11, color="FF0000", bold=True)
        red_bold = Font(bold=True, color="FF0000", name="Calibri", size=12)
        red_bold_big = Font(bold=True, color="FF0000", name="Calibri", size=13)
        # --- NEW FONTS ---
        red_bold_large = Font(bold=True, color="FF0000", name="Calibri", size=14)
        red_bold_extra_large = Font(bold=True, color="FF0000", name="Calibri", size=16)
        red_bold_note = Font(bold=True, color="FF0000", name="Calibri", size=11)
        
        bold_font = Font(bold=True, name='Calibri', size=11)
        italic_small = Font(italic=True, size=9, name='Calibri')
        
        # 4. Fills
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF66")
        grey_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        
        # --- NEW FILLS FOR PROFIT/PERCENTAGE ---
        light_green_fill = PatternFill(fill_type="solid", fgColor="E2EFDA") # Light Green for %
        light_blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")  # Light Blue for Profit
        
        START_ROW = 12
        actual_rows = 0

        # --- A. FILL DATA ROWS (Loop over ALL requirements) ---
        for idx, item in enumerate(requirements): 
            row = START_ROW + idx
            actual_rows += 1
            try:
                # Col 1: SL NO
                ws.cell(row=row, column=1).value = idx + 1

                # Col 2: DESCRIPTION (Rich Text)
                desc_text = str(item.get("Description", "") or "N/A")
                offering_text = str(item.get("Company Offering", "") or "")

                header1 = TextBlock(InlineFont(b=True, u="single", color="800080"), "Your Requirement:\n")
                body1 = TextBlock(InlineFont(color="000000"), f"{desc_text}\n\n")
                header2 = TextBlock(InlineFont(b=True, u="single", color="FF0000"), "We OFFER:\n")
                body2 = TextBlock(InlineFont(color="000000"), f"{offering_text}")

                cell_desc = ws.cell(row=row, column=2)
                cell_desc.value = CellRichText([header1, body1, header2, body2])

                # Col 3: BRAND
                ws.cell(row=row, column=3).value = item.get("Brand and model", "")

                # Col 5: DELIVERY (RED BOLD)
                cell_del = ws.cell(row=row, column=5, value="Ex stock, subject to prior sales.")
                cell_del.font = red_bold_font

                # Col 6: QTY (BOLD)
                qty_val = self._to_float(item.get("Quantity", 0))
                cell_qty = ws.cell(row=row, column=6)
                cell_qty.value = qty_val
                cell_qty.number_format = "#,##0.00"
                cell_qty.font = bold_font

                # Col 7: UNIT (BOLD)
                cell_unit = ws.cell(row=row, column=7)
                cell_unit.value = item.get("Unit", "")
                cell_unit.font = bold_font

                # Col 8: UNIT PRICE (BOLD)
                price_val = self._to_float(item.get("Unit price", 0))
                cell_price = ws.cell(row=row, column=8)
                cell_price.value = price_val 
                cell_price.number_format = "#,##0.00"
                cell_price.font = bold_font

                # Col 9: TOTAL PRICE FORMULA (BOLD)
                cell_total = ws.cell(row=row, column=9)
                cell_total.value = f"=F{row}*H{row}"
                cell_total.number_format = "#,##0.00"
                cell_total.font = bold_font
                
                # --- INTERNAL CALCULATIONS (CP/Profit) ---
                
                # Col K (11): CP Input (BOLD)
                cell_cp = ws.cell(row=row, column=11)
                cell_cp.value = 0.00
                cell_cp.number_format = "#,##0.00"
                cell_cp.font = bold_font

                # Col L (12): % Input (Light Green Background)
                cell_pct = ws.cell(row=row, column=12)
                cell_pct.value = 0.00
                cell_pct.number_format = "0.00%"
                cell_pct.fill = light_green_fill

                # Col M (13): Profit Formula (BOLD + Light Blue Background)
                cell_profit = ws.cell(row=row, column=13)
                cell_profit.value = f"=(N{row}-K{row})*F{row}"
                cell_profit.number_format = "#,##0.00"
                cell_profit.font = bold_font
                cell_profit.fill = light_blue_fill

                # Col N (14): SP Formula (BOLD)
                cell_sp = ws.cell(row=row, column=14)
                cell_sp.value = f"=K{row}*(1+L{row})"
                cell_sp.number_format = "#,##0.00"
                cell_sp.font = bold_font

                # --- APPLY BORDERS & ALIGNMENT ---
                ws.row_dimensions[row].height = 140 
                
                for col in range(1, 15):
                    cell = ws.cell(row=row, column=col)
                    cell.border = full_border 
                    
                    if col == 2:
                        cell.alignment = left_top_align
                    else:
                        cell.alignment = center_align

            except Exception as row_error:
                logger.error(f"Error processing row {row}: {row_error}")
                continue

        # --- B. TOTALS & FOOTER ---
        try:
            last_data_row = START_ROW + max(actual_rows, 1) - 1
            
            TOTAL_ROW = last_data_row + 1
            VAT_ROW = TOTAL_ROW + 1
            GRAND_ROW = VAT_ROW + 1
            NOTE_ROW = GRAND_ROW + 1  # Note row immediately follows Grand Total

            # ---- TOTAL AMOUNT ----
            ws.merge_cells(start_row=TOTAL_ROW, start_column=1, end_row=TOTAL_ROW, end_column=8)
            ws.cell(row=TOTAL_ROW, column=1, value="Total Amount (AED).")
            ws.cell(row=TOTAL_ROW, column=9, value=f"=SUM(I{START_ROW}:I{last_data_row})")
            
            # ---- VAT ----
            ws.merge_cells(start_row=VAT_ROW, start_column=1, end_row=VAT_ROW, end_column=8)
            ws.cell(row=VAT_ROW, column=1, value="VAT 5% (AED).")
            ws.cell(row=VAT_ROW, column=9, value=f"=I{TOTAL_ROW}*0.05")

            # ---- GRAND TOTAL ----
            ws.merge_cells(start_row=GRAND_ROW, start_column=1, end_row=GRAND_ROW, end_column=8)
            ws.cell(row=GRAND_ROW, column=1, value="GRAND TOTAL AMOUNT (AED).")
            ws.cell(row=GRAND_ROW, column=9, value=f"=SUM(I{TOTAL_ROW}:I{VAT_ROW})")

            # ---- NOTE ROW ----
            ws.merge_cells(start_row=NOTE_ROW, start_column=1, end_row=NOTE_ROW, end_column=9)
            ws.cell(row=NOTE_ROW, column=1, value="NOTE :- STOCK AVAILBILITY AS PER THE QUOTATION DATE KINDLY CONFIRM AT THE TIME OF CONFIRMATION.")
            
            # ---- APPLY FORMATTING ----
            
            # 1. Total Amount & VAT (14pt Red Bold)
            for r in (TOTAL_ROW, VAT_ROW):
                ws.row_dimensions[r].height = 20 # Increase height
                ws.cell(row=r, column=1).font = red_bold_large
                ws.cell(row=r, column=9).font = red_bold_large
                ws.cell(row=r, column=1).alignment = right_align
                ws.cell(row=r, column=9).alignment = center_align
                ws.cell(row=r, column=9).number_format = "#,##0.00"
                for c in range(1, 10):
                    ws.cell(row=r, column=c).border = full_border # Full thin border

            # 2. Grand Total (16pt Red Bold, Yellow Fill)
            ws.row_dimensions[GRAND_ROW].height = 25 # Increase height
            ws.cell(row=GRAND_ROW, column=1).font = red_bold_extra_large
            ws.cell(row=GRAND_ROW, column=9).font = red_bold_extra_large
            ws.cell(row=GRAND_ROW, column=1).alignment = right_align
            ws.cell(row=GRAND_ROW, column=9).alignment = center_align
            ws.cell(row=GRAND_ROW, column=9).number_format = "#,##0.00"
            for c in range(1, 10):
                cell = ws.cell(row=GRAND_ROW, column=c)
                cell.fill = yellow_fill
                cell.border = full_border # Full thin border

            # 3. Note Row (Red Bold Note)

            note_cell = ws.cell(row=NOTE_ROW, column=1)
            note_cell.font = red_bold_note
            note_cell.alignment = left_center_align
            for c in range(1, 10):
                ws.cell(row=NOTE_ROW, column=c).border = full_border

            # --- TERMS & CONDITIONS ---
            TERMS_ROW = NOTE_ROW + 1  # Start 2 rows after Note

            ws.merge_cells(start_row=TERMS_ROW, start_column=1, end_row=TERMS_ROW, end_column=9)
            t_cell = ws.cell(row=TERMS_ROW, column=1, value="TERMS:")
            t_cell.font = bold_font
            t_cell.fill = grey_fill
            t_cell.alignment = left_center_align
            for c in range(1, 10):
                ws.cell(row=TERMS_ROW, column=c).border = full_border

            # ---- TERMS ROWS ----
            terms_data = [
                ("Price:", "Ex Warehouse Dubai, Packed."),
                ("Delivery:", "Stated in Description Column Against Each Item."),
                ("Payment:", "30 Days Credit."),
                ("Validity:", "15 days from offer date.")
            ]

            # Fix: Track current_row for footer logic
            current_row = TERMS_ROW
            for i, (label, value) in enumerate(terms_data):
                current_row = TERMS_ROW + 1 + i
                ws.cell(row=current_row, column=1, value=label)
                ws.cell(row=current_row, column=1).font = bold_font
                ws.cell(row=current_row, column=1).alignment = left_center_align

                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
                ws.cell(row=current_row, column=2, value=value)
                ws.cell(row=current_row, column=2).font = bold_font
                ws.cell(row=current_row, column=2).alignment = left_center_align

                for c in range(1, 10):
                    ws.cell(row=current_row, column=c).border = full_border

            # --- FOOTER MESSAGES & DYNAMIC PRINT AREA ---
            # --- FOOTER MESSAGES & DYNAMIC PRINT AREA ---
            # Define rows for footer structure
            # current_row (Terms End) -> Spacer (1) -> Msg -> Spacer (1) -> Regards -> Company -> Spacer (2) -> Disclaimer
            
            FS_START = current_row + 1
            MSG_ROW = FS_START + 1
            REGARDS_ROW = MSG_ROW + 2
            COMPANY_ROW = REGARDS_ROW + 1
            DISCLAIMER_ROW = COMPANY_ROW + 3
            FS_END = DISCLAIMER_ROW
            
            # Apply borders to the entire block (Outline only, no inner grid)
            for r in range(FS_START, FS_END + 1):
                # Merge row first
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                
                for c in range(1, 10):
                    # Determine borders based on position
                    b_left = thin_side if c == 1 else None
                    b_right = thin_side if c == 9 else None
                    b_bottom = thin_side if r == FS_END else None
                    # Top border is handled by the row above (Terms), so None here to avoid double/inner lines
                    
                    cell_border = Border(left=b_left, right=b_right, top=None, bottom=b_bottom)
                    ws.cell(row=r, column=c).border = cell_border
            
            # 1. Message
            ws.cell(row=MSG_ROW, column=1, value="Please revert for clarifications if any. Thank you for providing an opportunity to quote.")
            ws.cell(row=MSG_ROW, column=1).alignment = left_center_align
            
            # 2. Best Regards
            ws.cell(row=REGARDS_ROW, column=1, value="Best Regards,")
            ws.cell(row=REGARDS_ROW, column=1).alignment = left_center_align
            
            # 3. Company Name (Bold)
            ws.cell(row=COMPANY_ROW, column=1, value="Dbest Building Hardware and Tools Trading LLC.")
            ws.cell(row=COMPANY_ROW, column=1).font = bold_font
            ws.cell(row=COMPANY_ROW, column=1).alignment = left_center_align
        
            # 4. Disclaimer
            disc_cell = ws.cell(row=DISCLAIMER_ROW, column=1, value="(This message has been electronically transmitted and does not require a signature).")
            disc_cell.font = italic_small
            disc_cell.alignment = left_center_align

            # THE FIX: Extend Blue Line to include template's contact bar images
            FINAL_PRINT_ROW = DISCLAIMER_ROW + 15 
            ws.print_area = f'A1:I{FINAL_PRINT_ROW}'

            # Force scaling to fit content
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0 

        except Exception as e:
            logger.error(f"Error writing totals/footer: {e}")
            import traceback
            logger.error(traceback.format_exc())

    def _to_float(self, value) -> float:
        try:
            if value is None: return 0.0
            s = str(value).strip().replace(",", "")
            return float(s) if s else 0.0
        except: 
            return 0.0